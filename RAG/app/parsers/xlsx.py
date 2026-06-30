"""XLSX parser using openpyxl, sheet-by-sheet processing, formulas, and BOQ awareness."""
from __future__ import annotations

import io
import logging
from typing import Any
from app.parsers import ParserOutput, register_parser

logger = logging.getLogger(__name__)

@register_parser(extensions=[".xlsx", ".xls"], mime_types=["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"])
def parse_xlsx(filename: str, content: bytes, content_type: str) -> ParserOutput:
    """Parse XLSX files, handle multiple sheets, output cell contents, resolve mergers, flag formulas, detect BOQ structures."""
    try:
        import openpyxl
    except ModuleNotFoundError as e:
        logger.error("openpyxl is not installed in the RAG environment.")
        raise RuntimeError("openpyxl is required to parse Excel spreadsheets.") from e

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=False)
    except Exception as e:
        logger.error(f"Failed to open Excel workbook: {e}")
        raise ValueError(f"Unable to read Excel file: {e}") from e

    output_lines = [f"Excel File: {filename}", f"Sheets: {', '.join(wb.sheetnames)}"]
    has_boq = False
    boq_headers = {"description", "quantity", "unit", "rate", "price", "amount", "total", "item no", "spec", "specification"}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        output_lines.append(f"\n--- Sheet: {sheet_name} ---")

        # Resolve merged cell mappings so child cells read parent's value
        merged_cell_map = {}
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
            parent_val = sheet.cell(row=min_row, column=min_col).value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    merged_cell_map[(r, c)] = parent_val

        # Read sheet row-by-row
        for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
            row_vals = []
            for col_idx, cell in enumerate(row, start=1):
                val = cell.value
                # If merged cell, use parent cell value
                if val is None and (row_idx, col_idx) in merged_cell_map:
                    val = merged_cell_map[(row_idx, col_idx)]

                if val is not None:
                    # check if it is a formula
                    if isinstance(val, str) and val.startswith("="):
                        row_vals.append(f"[Formula: {val}]")
                    else:
                        row_vals.append(str(val))
                else:
                    row_vals.append("")

            # Clean row values
            row_str = " | ".join(cell_val.strip() for cell_val in row_vals if cell_val.strip())
            if row_str:
                output_lines.append(f"Row {row_idx}: {row_str}")

            # Simple BOQ heuristic: Check if any row contains multiple BOQ keywords
            row_lower = [v.lower() for v in row_vals if isinstance(v, str)]
            boq_keywords_found = sum(1 for kw in boq_headers if any(kw in cell_val for cell_val in row_lower))
            if boq_keywords_found >= 3:
                has_boq = True

    full_content = "\n".join(output_lines)

    metadata = {
        "file_type": "xlsx",
        "page_count": len(wb.sheetnames),
        "has_drawings": False,
        "has_tables": True,
        "language": "en",
        "detected_domain": "architecture-civil" if has_boq else "general",
        "confidence": 1.0,
        "filename": filename,
        "is_boq": has_boq,
    }

    return ParserOutput(content=full_content, metadata=metadata)
